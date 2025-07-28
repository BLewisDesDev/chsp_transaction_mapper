from pathlib import Path
from typing import Dict, List, Any
from datetime import datetime, date
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from models.reconciliation_report import ReconciliationReport
from models.transaction import Transaction
from core.client_map_loader import ClientMapLoader


class ShiftCareReport:
    """Creates comprehensive Excel reports for ShiftCare invoice reconciliation."""
    
    def __init__(self, client_map: ClientMapLoader):
        self.client_map = client_map
        
    def export_excel_report(self, 
                           invoices_data: List[Dict],
                           shifts_data: List[Dict],
                           report: ReconciliationReport,
                           output_path: str) -> str:
        """
        Export ShiftCare reconciliation report to Excel with 3 sheets.
        
        Args:
            invoices_data: List of invoice dictionaries
            shifts_data: List of shift dictionaries  
            report: Reconciliation report with match results
            output_path: Path to save Excel file
            
        Returns:
            Path to created Excel file
        """
        # Create output directory if it doesn't exist
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        
        # Create workbook
        wb = Workbook()
        
        # Create the three sheets
        self._create_invoices_sheet(wb, invoices_data, report)
        self._create_client_summary_sheet(wb, invoices_data, shifts_data, report)
        self._create_shifts_sheet(wb, shifts_data, report)
        
        # Save the workbook
        wb.save(output_file)
        return str(output_file)
    
    def _create_invoices_sheet(self, 
                              wb: Workbook, 
                              invoices_data: List[Dict],
                              report: ReconciliationReport):
        """Create the invoices sheet with matched/unmatched highlighting."""
        # Remove default sheet and create invoices sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        ws = wb.create_sheet("Invoices", 0)
        
        # Create match status lookup
        match_status = {}
        for match_result in report.match_results:
            match_status[match_result.transaction_id] = match_result.is_matched
        
        # Headers
        headers = [
            "Matched", "Invoice ID", "Invoice Number", "Client ID", "Client Name", 
            "Total Amount", "Status", "Due Date", "Issued Date", "Created Date",
            "Caura Client ID", "ACN", "DEX DA ID", "DEX HM ID"
        ]
        ws.append(headers)
        
        # Add invoice data rows
        for invoice in invoices_data:
            invoice_id = str(invoice.get('id', ''))
            transaction_id = f"invoice_{invoice_id}"
            is_matched = match_status.get(transaction_id, False)
            
            # Get client details from client map
            client_details = self._get_client_details_for_invoice(invoice)
            
            row = [
                'Matched' if is_matched else 'Unmatched',
                invoice_id,
                invoice.get('reference_number', ''),
                invoice.get('client_id', ''),
                client_details.get('client_name', ''),
                f"${float(invoice.get('total_amount', 0)):.2f}",
                invoice.get('status', ''),
                invoice.get('due_at', ''),
                invoice.get('issued_at', ''),
                invoice.get('created_at', ''),
                client_details.get('caura_id', ''),
                client_details.get('acn', ''),
                client_details.get('dex_da_id', ''),
                client_details.get('dex_hm_id', '')
            ]
            ws.append(row)
        
        # Apply formatting
        self._format_sheet_headers(ws)
        self._apply_matched_coloring(ws, match_status, 'A')
    
    def _create_client_summary_sheet(self, 
                                   wb: Workbook, 
                                   invoices_data: List[Dict],
                                   shifts_data: List[Dict], 
                                   report: ReconciliationReport):
        """Create client summary sheet with service totals and identifiers."""
        ws = wb.create_sheet("Client Summary")
        
        # Headers
        headers = [
            "Client Name", "Caura ID", "ShiftCare Client ID", "ACN", "Email",
            "DEX DA Client ID", "DEX HM Client ID", "DEX DA Case ID", "DEX HM Case ID", "SLK",
            "Service Types", "Total Amount", "Invoice Count", "Shift Count",
            "First Service Date", "Last Service Date", "Matched Status"
        ]
        ws.append(headers)
        
        # Calculate client summaries
        client_summaries = self._calculate_client_summaries(invoices_data, shifts_data, report)
        
        # Add client data rows
        for client_id, summary in client_summaries.items():
            row = [
                summary.get('client_name', ''),
                summary.get('caura_id', ''),
                client_id,
                summary.get('acn', ''),
                summary.get('email', ''),
                summary.get('dex_da_client_id', ''),
                summary.get('dex_hm_client_id', ''),
                summary.get('dex_da_case_id', ''),
                summary.get('dex_hm_case_id', ''),
                summary.get('slk', ''),
                ', '.join(summary.get('service_types', [])),
                f"${summary.get('total_amount', 0):.2f}",
                summary.get('invoice_count', 0),
                summary.get('shift_count', 0),
                summary.get('first_service_date', ''),
                summary.get('last_service_date', ''),
                'Matched' if summary.get('is_matched', False) else 'Unmatched'
            ]
            ws.append(row)
        
        # Format headers
        self._format_sheet_headers(ws)
    
    def _create_shifts_sheet(self, 
                           wb: Workbook, 
                           shifts_data: List[Dict],
                           report: ReconciliationReport):
        """Create shifts sheet with individual shift details and client identifiers."""
        ws = wb.create_sheet("Shifts")
        
        # Headers
        headers = [
            "Shift ID", "Invoice ID", "Client ID", "Client Name", "Caura ID", 
            "Service Date", "Service Type", "Hours", "Rate", "Amount", 
            "Description", "ACN", "DEX DA ID", "DEX HM ID", "SLK", "Paid Status"
        ]
        ws.append(headers)
        
        # Add shift data rows
        for shift in shifts_data:
            # Get client details
            client_details = self._get_client_details_by_shiftcare_id(shift.get('client_id'))
            
            # Extract service date from description
            service_date = self._extract_service_date_from_description(shift.get('description', ''))
            
            row = [
                shift.get('shift_id', ''),
                shift.get('invoice_id', ''),
                shift.get('client_id', ''),
                client_details.get('client_name', ''),
                client_details.get('caura_id', ''),
                service_date,
                shift.get('pricebook_name', ''),
                shift.get('quantity', 0),
                f"${float(shift.get('rate', 0)):.2f}",
                f"${float(shift.get('amount', 0)):.2f}",
                shift.get('description', ''),
                client_details.get('acn', ''),
                client_details.get('dex_da_id', ''),
                client_details.get('dex_hm_id', ''),
                client_details.get('slk', ''),
                'Paid'  # All shifts in this report are from paid invoices
            ]
            ws.append(row)
        
        # Format headers
        self._format_sheet_headers(ws)
    
    def _get_client_details_for_invoice(self, invoice: Dict) -> Dict:
        """Get client details from client map for an invoice."""
        client_id = str(invoice.get('client_id', ''))
        return self._get_client_details_by_shiftcare_id(client_id)
    
    def _get_client_details_by_shiftcare_id(self, shiftcare_client_id: str) -> Dict:
        """Get client details from client map by ShiftCare client ID."""
        if not shiftcare_client_id:
            return {}
            
        self.client_map.load_client_map()
        
        # Search for client by ShiftCare platform identifier
        for caura_id, client in self.client_map._client_cache.items():
            platform_identifiers = client.get("platform_identifiers", [])
            
            for platform_id in platform_identifiers:
                if platform_id.get("platform") in ["shiftcare_domestic_assistance", "shiftcare_home_maintenance"]:
                    if platform_id.get("identifiers", {}).get("client_id") == shiftcare_client_id:
                        return self._extract_client_details(client, caura_id)
        
        return {'client_id': shiftcare_client_id}
    
    def _extract_client_details(self, client: Dict, caura_id: str) -> Dict:
        """Extract comprehensive client details."""
        personal_info = client.get("personal_info", {})
        
        # Get name
        given_name = personal_info.get("given_name", "")
        family_name = personal_info.get("family_name", "")
        client_name = f"{given_name} {family_name}".strip()
        
        # Get email
        emails = personal_info.get("emails", [])
        email = emails[0] if emails else ""
        
        # Get platform identifiers
        platform_ids = self._get_platform_identifiers(client)
        
        return {
            'caura_id': caura_id,
            'client_name': client_name,
            'email': email,
            'acn': platform_ids.get('acn', ''),
            'dex_da_client_id': platform_ids.get('dex_da_client_id', ''),
            'dex_hm_client_id': platform_ids.get('dex_hm_client_id', ''),
            'dex_da_case_id': platform_ids.get('dex_da_case_id', ''),
            'dex_hm_case_id': platform_ids.get('dex_hm_case_id', ''),
            'slk': platform_ids.get('slk', '')
        }
    
    def _get_platform_identifiers(self, client: Dict) -> Dict[str, str]:
        """Extract platform identifiers from client."""
        platform_ids = {}
        
        for platform_id in client.get("platform_identifiers", []):
            platform = platform_id.get("platform", "")
            identifiers = platform_id.get("identifiers", {})
            
            if platform == "aged_care":
                platform_ids['acn'] = identifiers.get("acn", "")
            elif platform == "dex_da":
                platform_ids['dex_da_client_id'] = identifiers.get("client_id", "")
                platform_ids['dex_da_case_id'] = identifiers.get("case_id", "")
                platform_ids['slk'] = identifiers.get("slk", "")
            elif platform == "dex_hm":
                platform_ids['dex_hm_client_id'] = identifiers.get("client_id", "")
                platform_ids['dex_hm_case_id'] = identifiers.get("case_id", "")
                # SLK should be the same across platforms
                if not platform_ids.get('slk'):
                    platform_ids['slk'] = identifiers.get("slk", "")
        
        return platform_ids
    
    def _calculate_client_summaries(self, 
                                  invoices_data: List[Dict], 
                                  shifts_data: List[Dict],
                                  report: ReconciliationReport) -> Dict:
        """Calculate client summaries with service dates and totals."""
        summaries = {}
        
        # Create match status lookup
        match_status = {}
        for match_result in report.match_results:
            match_status[match_result.transaction_id] = match_result.is_matched
        
        # Process invoices for totals
        for invoice in invoices_data:
            client_id = str(invoice.get('client_id', ''))
            if not client_id:
                continue
                
            if client_id not in summaries:
                client_details = self._get_client_details_by_shiftcare_id(client_id)
                summaries[client_id] = {
                    **client_details,
                    'total_amount': 0,
                    'invoice_count': 0,
                    'shift_count': 0,
                    'service_types': set(),
                    'service_dates': [],
                    'is_matched': False
                }
            
            summaries[client_id]['total_amount'] += float(invoice.get('total_amount', 0))
            summaries[client_id]['invoice_count'] += 1
            
            # Check if this invoice is matched
            transaction_id = f"invoice_{invoice.get('id', '')}"
            if match_status.get(transaction_id, False):
                summaries[client_id]['is_matched'] = True
        
        # Process shifts for service details
        for shift in shifts_data:
            client_id = str(shift.get('client_id', ''))
            if client_id in summaries:
                summaries[client_id]['shift_count'] += 1
                
                # Add service type
                service_type = shift.get('pricebook_name', '')
                if service_type:
                    summaries[client_id]['service_types'].add(service_type)
                
                # Extract service date
                service_date = self._extract_service_date_from_description(shift.get('description', ''))
                if service_date:
                    summaries[client_id]['service_dates'].append(service_date)
        
        # Calculate first and last service dates
        for client_id, summary in summaries.items():
            service_dates = summary['service_dates']
            if service_dates:
                summary['first_service_date'] = min(service_dates)
                summary['last_service_date'] = max(service_dates)
            else:
                summary['first_service_date'] = ''
                summary['last_service_date'] = ''
            
            # Convert service_types set to list
            summary['service_types'] = list(summary['service_types'])
            
            # Remove the temporary service_dates list
            del summary['service_dates']
        
        return summaries
    
    def _extract_service_date_from_description(self, description: str) -> str:
        """Extract service date from shift description."""
        import re
        
        # Look for date pattern like "03/02/2025" in the description
        date_pattern = r'(\d{2}/\d{2}/\d{4})'
        match = re.search(date_pattern, description)
        
        if match:
            return match.group(1)
        
        return ''
    
    def _format_sheet_headers(self, ws):
        """Apply formatting to sheet headers."""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Blue
        header_font = Font(bold=True, color="FFFFFF")  # White text
        
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            
            # Auto-adjust column width
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].auto_size = True
    
    def _apply_matched_coloring(self, ws, match_status: Dict, column_letter: str):
        """Apply matched/unmatched coloring to a specific column."""
        matched_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
        unmatched_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red
        
        # Start from row 2 (after header)
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)  # Column A (Matched column)
            if cell.value == 'Matched':
                cell.fill = matched_fill
            else:
                cell.fill = unmatched_fill