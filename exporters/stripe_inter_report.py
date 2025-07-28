from pathlib import Path
from typing import Dict, List, Any
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from models.reconciliation_report import ReconciliationReport
from models.transaction import Transaction
from core.client_map_loader import ClientMapLoader


class StripeInterReport:
    """Creates Excel reports for Stripe transaction reconciliation with matched/unmatched highlighting."""
    
    def __init__(self, client_map: ClientMapLoader):
        self.client_map = client_map
        
    def export_excel_report(self, 
                           transactions_df: pd.DataFrame,
                           transactions: List[Transaction],
                           report: ReconciliationReport,
                           output_path: str) -> str:
        """
        Export Stripe reconciliation report to Excel with highlighting and summary sheets.
        
        Args:
            transactions_df: Original transaction dataframe
            transactions: List of Transaction objects
            report: Reconciliation report with match results
            output_path: Path to save Excel file
            
        Returns:
            Path to created Excel file
        """
        # Create output directory if it doesn't exist
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        
        # Create workbook
        wb = Workbook()
        
        # Create transactions sheet with highlighting
        self._create_transactions_sheet(wb, transactions_df, report)
        
        # Create summary sheets
        self._create_client_summary_sheet(wb, transactions, report)
        self._create_unmatched_emails_sheet(wb, transactions, report)
        
        # Save the workbook
        wb.save(output_file)
        return str(output_file)
    
    def _create_transactions_sheet(self, 
                                 wb: Workbook, 
                                 transactions_df: pd.DataFrame,
                                 report: ReconciliationReport):
        """Create the main transactions sheet with matched/unmatched highlighting."""
        # Remove default sheet and create transactions sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        ws = wb.create_sheet("Transactions", 0)
        
        # Create match status lookup
        match_status = {}
        for match_result in report.match_results:
            match_status[match_result.transaction_id] = match_result.is_matched
        
        # Add or update Matched column 
        enhanced_df = transactions_df.copy()
        
        # Determine ID column (different for different platforms)
        id_column = None
        if 'id' in enhanced_df.columns:
            id_column = 'id'
        elif enhanced_df.index.name:
            id_column = enhanced_df.index.name
        else:
            # For paper receipts, create an ID column based on row index
            enhanced_df['transaction_id'] = [f"receipt_{i+1}" for i in range(len(enhanced_df))]
            id_column = 'transaction_id'
        
        # Check if Matched column already exists
        if 'Matched' not in enhanced_df.columns:
            if id_column == 'transaction_id':
                enhanced_df.insert(0, 'Matched', enhanced_df[id_column].map(
                    lambda x: 'Matched' if match_status.get(str(x), False) else 'Unmatched'
                ))
            else:
                enhanced_df.insert(0, 'Matched', enhanced_df[id_column].map(
                    lambda x: 'Matched' if match_status.get(str(x), False) else 'Unmatched'
                ))
        else:
            # Update existing Matched column
            enhanced_df['Matched'] = enhanced_df[id_column].map(
                lambda x: 'Matched' if match_status.get(str(x), False) else 'Unmatched'
            )
        
        # Add dataframe to worksheet
        for r in dataframe_to_rows(enhanced_df, index=False, header=True):
            ws.append(r)
        
        # Apply formatting
        self._format_transactions_sheet(ws, match_status, enhanced_df)
    
    def _format_transactions_sheet(self, ws, match_status: Dict[str, bool], df: pd.DataFrame):
        """Apply color formatting to the transactions sheet."""
        # Define fill colors
        matched_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
        unmatched_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light gray
        header_font = Font(bold=True)
        
        # Format headers
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
        
        # Format Matched column (column A) based on match status
        matched_col = 1  # Column A
        for row in range(2, len(df) + 2):  # Start from row 2 (after header)
            cell = ws.cell(row=row, column=matched_col)
            if cell.value == 'Matched':
                cell.fill = matched_fill
            else:
                cell.fill = unmatched_fill
        
        # Auto-adjust column widths
        for col in range(1, len(df.columns) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].auto_size = True
    
    def _create_client_summary_sheet(self, wb: Workbook, transactions: List[Transaction], report: ReconciliationReport):
        """Create summary sheet with matched client totals and details."""
        ws = wb.create_sheet("Client Summary")
        
        # Headers
        headers = [
            "Client Name", "Email", "ACN", "Service Types", "Total Amount", 
            "Transaction Count", "DEX DA Client ID", "DEX HM Client ID", 
            "DEX DA Case ID", "DEX HM Case ID", "SLK"
        ]
        ws.append(headers)
        
        # Collect client data
        client_totals = self._calculate_client_totals(transactions, report)
        
        # Add client data rows
        for caura_id, data in client_totals.items():
            client = self.client_map.get_client(caura_id)
            if not client:
                continue
                
            # Extract client information
            personal_info = client.get("personal_info", {})
            client_name = f"{personal_info.get('given_name', '')} {personal_info.get('family_name', '')}".strip()
            emails = ", ".join(personal_info.get("emails", []))
            
            # Extract ACN
            acn = self._get_acn_from_client(client)
            
            # Extract service types
            service_types = self._get_service_types_from_client(client)
            
            # Extract platform identifiers
            platform_ids = self._get_platform_identifiers(client)
            
            row = [
                client_name,
                emails,
                acn,
                ", ".join(service_types),
                f"${data['total_amount']:.2f}",
                data['transaction_count'],
                platform_ids.get('dex_da_client_id', ''),
                platform_ids.get('dex_hm_client_id', ''),
                platform_ids.get('dex_da_case_id', ''),
                platform_ids.get('dex_hm_case_id', ''),
                platform_ids.get('slk', '')
            ]
            ws.append(row)
        
        # Format headers
        self._format_summary_headers(ws)
    
    def _create_unmatched_emails_sheet(self, wb: Workbook, transactions: List[Transaction], report: ReconciliationReport):
        """Create sheet with unmatched email addresses."""
        ws = wb.create_sheet("Unmatched Emails")
        
        # Headers
        headers = ["Email", "Transaction Count", "Total Amount", "Sample Transaction IDs"]
        ws.append(headers)
        
        # Collect unmatched email data
        unmatched_emails = self._collect_unmatched_emails(transactions, report)
        
        # Add unmatched email rows
        for email, data in unmatched_emails.items():
            sample_ids = ", ".join(data['transaction_ids'][:3])  # Show first 3 transaction IDs
            if len(data['transaction_ids']) > 3:
                sample_ids += f" (+{len(data['transaction_ids']) - 3} more)"
                
            row = [
                email,
                data['count'],
                f"${data['total_amount']:.2f}",
                sample_ids
            ]
            ws.append(row)
        
        # Format headers
        self._format_summary_headers(ws)
    
    def _calculate_client_totals(self, transactions: List[Transaction], report: ReconciliationReport) -> Dict[str, Dict]:
        """Calculate totals per matched client."""
        client_totals = {}
        
        # Create lookup for match results by transaction ID
        match_lookup = {mr.transaction_id: mr for mr in report.match_results}
        
        for transaction in transactions:
            match_result = match_lookup.get(transaction.transaction_id)
            if match_result and match_result.is_matched and match_result.client_caura_id:
                caura_id = match_result.client_caura_id
                
                if caura_id not in client_totals:
                    client_totals[caura_id] = {
                        'total_amount': 0,
                        'transaction_count': 0
                    }
                
                client_totals[caura_id]['total_amount'] += float(transaction.amount)
                client_totals[caura_id]['transaction_count'] += 1
        
        return client_totals
    
    def _collect_unmatched_emails(self, transactions: List[Transaction], report: ReconciliationReport) -> Dict[str, Dict]:
        """Collect unmatched email addresses and their transaction details."""
        unmatched_emails = {}
        
        # Create lookup for match results by transaction ID
        match_lookup = {mr.transaction_id: mr for mr in report.match_results}
        
        for transaction in transactions:
            match_result = match_lookup.get(transaction.transaction_id)
            if match_result and not match_result.is_matched and transaction.email:
                email = transaction.email
                
                if email not in unmatched_emails:
                    unmatched_emails[email] = {
                        'count': 0,
                        'total_amount': 0,
                        'transaction_ids': []
                    }
                
                unmatched_emails[email]['count'] += 1
                unmatched_emails[email]['total_amount'] += float(transaction.amount)
                unmatched_emails[email]['transaction_ids'].append(transaction.transaction_id)
        
        return unmatched_emails
    
    def _get_acn_from_client(self, client: Dict) -> str:
        """Extract ACN from client platform identifiers."""
        for platform_id in client.get("platform_identifiers", []):
            if platform_id.get("platform") == "aged_care":
                return platform_id.get("identifiers", {}).get("acn", "")
        return ""
    
    def _get_service_types_from_client(self, client: Dict) -> List[str]:
        """Extract service types from client service information."""
        service_types = []
        services = client.get("service_information", {}).get("services", [])
        for service in services:
            service_type = service.get("service_type", "")
            if service_type:
                service_types.append(service_type)
        return service_types
    
    def _get_platform_identifiers(self, client: Dict) -> Dict[str, str]:
        """Extract platform identifiers from client."""
        platform_ids = {}
        
        for platform_id in client.get("platform_identifiers", []):
            platform = platform_id.get("platform", "")
            identifiers = platform_id.get("identifiers", {})
            
            if platform == "dex_da":
                platform_ids['dex_da_client_id'] = identifiers.get("client_id", "")
                platform_ids['dex_da_case_id'] = identifiers.get("case_id", "")
                platform_ids['slk'] = identifiers.get("slk", "")
            elif platform == "dex_hm":
                platform_ids['dex_hm_client_id'] = identifiers.get("client_id", "")
                platform_ids['dex_hm_case_id'] = identifiers.get("case_id", "")
                # SLK should be the same across platforms, but we'll use it from either
                if not platform_ids.get('slk'):
                    platform_ids['slk'] = identifiers.get("slk", "")
        
        return platform_ids
    
    def _format_summary_headers(self, ws):
        """Apply formatting to summary sheet headers."""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Blue
        header_font = Font(bold=True, color="FFFFFF")  # White text
        
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            
            # Auto-adjust column width
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].auto_size = True